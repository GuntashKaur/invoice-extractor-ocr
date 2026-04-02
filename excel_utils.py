import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

EXCEL_FILE = "invoice_data.xlsx"


# -----------------------------
# SAVE DATA (VERTICAL + APPEND)
# -----------------------------
def save_to_excel(data):

    # Convert dictionary → vertical format
    df_new = pd.DataFrame(list(data.items()), columns=["Field", "Value"])

    # Add separator for new invoice (clean look)
    separator = pd.DataFrame([["---- NEW INVOICE ----", ""]], columns=["Field", "Value"])

    if os.path.exists(EXCEL_FILE):
        df_existing = pd.read_excel(EXCEL_FILE)
        df_final = pd.concat([df_existing, separator, df_new], ignore_index=True)
    else:
        df_final = df_new

    # Save to Excel
    df_final.to_excel(EXCEL_FILE, index=False)

    # -----------------------------
    # FORMATTING (IMPORTANT)
    # -----------------------------
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 65

    # Apply wrap + alignment
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    wb.save(EXCEL_FILE)

    return EXCEL_FILE


# -----------------------------
# LOAD EXCEL (PREVIEW)
# -----------------------------
def load_excel():
    if os.path.exists(EXCEL_FILE):
        return pd.read_excel(EXCEL_FILE)
    return pd.DataFrame(columns=["Field", "Value"])


# -----------------------------
# RESET EXCEL
# -----------------------------
def reset_excel():
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)